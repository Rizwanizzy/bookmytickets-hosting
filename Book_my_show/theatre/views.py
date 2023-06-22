from django.shortcuts import render, redirect
from django.contrib.auth import authenticate, login, logout
from django.contrib import auth
from django.contrib import messages
from django.contrib.auth.models import User
from .models import *
from django.contrib.auth.decorators import login_required
from admin_dashboard.models import *
from home.models import UserProfile
from users.models import *
from django.http import JsonResponse
from Book_my_show.settings import KEY, SECRET
from home.forms import *
import razorpay, json
from decimal import Decimal
from datetime import date, timedelta, datetime
from django.db.models import Sum
from django.utils import timezone
import decimal
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
import openpyxl
from django.core.paginator import Paginator

# Create your views here.


def theatre_home(request):
    if "user" in request.session:
        return redirect("home")
    if "admin" in request.session:
        return redirect("admin_home")
    if "theatre" in request.session:
        # Get today's date
        today = date.today()

        # Get the start and end dates for the current month
        start_of_month = today.replace(day=1)
        end_of_month = start_of_month.replace(
            month=start_of_month.month + 1
        ) - timedelta(days=1)

        # Calculate the total sale price for today
        today_total_price = (
            Theatre_Sale_Report.objects.filter(
                name=request.user,
                date_added__year=today.year,
                date_added__month=today.month,
                date_added__day=today.day,
            ).aggregate(total_price=Sum("booking__price"))["total_price"]
            or 0.00
        )

        today_total_revenue = (
            Theatre_Sale_Report.objects.filter(
                name=request.user,
                date_added__year=today.year,
                date_added__month=today.month,
                date_added__day=today.day,
            ).aggregate(total_price=Sum("theatre_earnings"))["total_price"]
            or 0.00
        )

        # Calculate the sales and revenue data for each day of the week
        sales_data = []
        revenue_data = []
        days_of_week = []

        for i in range(7):
            # Get the date for the corresponding day of the week
            day = today - timedelta(days=today.weekday()) + timedelta(days=i)
            days_of_week.append(day.strftime("%a"))

            # Calculate the total sale price for the day
            day_total_price = (
                Theatre_Sale_Report.objects.filter(
                    name=request.user,
                    date_added__year=day.year,
                    date_added__month=day.month,
                    date_added__day=day.day,
                ).aggregate(total_price=Sum("booking__price"))["total_price"]
                or 0.00
            )
            sales_data.append(day_total_price)

            # Calculate the total revenue for the day
            day_total_revenue = (
                Theatre_Sale_Report.objects.filter(
                    name=request.user,
                    date_added__year=day.year,
                    date_added__month=day.month,
                    date_added__day=day.day,
                ).aggregate(total_price=Sum("theatre_earnings"))["total_price"]
                or 0.00
            )
            revenue_data.append(day_total_revenue)

        sale_reports = Theatre_Sale_Report.objects.filter(name=request.user)

        # Calculate the total sale price for this month
        month_total_price = (
            Theatre_Sale_Report.objects.filter(
                name=request.user, date_added__range=(start_of_month, end_of_month)
            ).aggregate(total_price=Sum("booking__price"))["total_price"]
            or 0.00
        )

        month_total_revenue = (
            Theatre_Sale_Report.objects.filter(
                name=request.user, date_added__range=(start_of_month, end_of_month)
            ).aggregate(total_price=Sum("theatre_earnings"))["total_price"]
            or 0.00
        )
        print(
            "sales_data:",
            sales_data,
            "revenue_data:",
            revenue_data,
            "days_of_week:",
            days_of_week,
        )
        revenue_data = [
            float(value) if isinstance(value, decimal.Decimal) else value
            for value in revenue_data
        ]
        context = {
            "sale_reports": sale_reports,
            "today_total_price": today_total_price,
            "month_total_price": month_total_price,
            "today_total_revenue": today_total_revenue,
            "month_total_revenue": month_total_revenue,
            "sales_data": sales_data,
            "revenue_data": revenue_data,
            "days_of_week": days_of_week,
        }
        return render(request, "theatre/theatre_home.html", context)
    else:
        return redirect("home")


def theatre_movies(request):
    if "user" in request.session:
        return redirect("home")
    if "admin" in request.session:
        return redirect("admin_home")
    if "theatre" in request.session:
        movies = Movies.objects.all()
        items_per_page = 4
        paginator = Paginator(movies, items_per_page)
        page_number = request.GET.get("page")
        page = paginator.get_page(page_number)
        return render(
            request, "theatre/theatre_movies.html", {"movies": movies, "page": page}
        )
    else:
        return redirect("home")


def theatre_screen(request):
    if "user" in request.session:
        return redirect("home")
    if "admin" in request.session:
        return redirect("admin_home")
    if "theatre" in request.session:
        # screens=Screen.objects.filter(theatre_id=request.user.id)
        user_profile = UserProfile.objects.get(user=request.user)
        screens = Screen.objects.filter(theatre=user_profile)
        items_per_page = 4
        paginator = Paginator(screens, items_per_page)
        page_number = request.GET.get("page")
        page = paginator.get_page(page_number)
        print("theatre_id", request.user.id)
        return render(
            request, "theatre/theatre_screen.html", {"screens": screens, "page": page}
        )
    else:
        return redirect("home")


def add_screen(request):
    if "user" in request.session:
        return redirect("home")
    if "admin" in request.session:
        return redirect("admin_home")
    if "theatre" in request.session:
        if request.method == "POST":
            # print('logged in username:',request.user)

            name = request.POST.get("name")
            price1 = request.POST.get("price1")
            price2 = request.POST.get("price2") or None
            price3 = request.POST.get("price3") or None
            movies_id = request.POST.get("movie")
            rows = request.POST.get("rows")
            columns = request.POST.get("columns")
            show_times = request.POST.getlist("show_times[]")

            theatre = UserProfile.objects.get(user=request.user)
            movies = Movies.objects.get(id=movies_id)
            screen = Screen.objects.create(
                theatre=theatre,
                name=name,
                price1=price1,
                price2=price2,
                price3=price3,
                movies=movies,
                total_seat_rows=rows,
                total_seat_columns=columns,
            )
            for show_time_str in show_times:
                show_time_obj = Show_Time.objects.create(time=show_time_str)
                screen.show_times.add(show_time_obj)
            screen.save()
            return redirect("theatre_screen")
        else:
            movies = Movies.objects.all()
            # shows=Show_Time.objects.all()
            return render(request, "theatre/add_screen.html", {"movies": movies})
    else:
        return redirect("home")


def update_screen(request, id):
    if "user" in request.session:
        return redirect("home")
    if "admin" in request.session:
        return redirect("admin_home")
    if "theatre" in request.session:
        screen = Screen.objects.get(id=id)
        if request.method == "POST":
            theatre = UserProfile.objects.get(user=request.user)
            name = request.POST.get("name")
            price1 = request.POST.get("price1")
            price2 = request.POST.get("price2") or None
            price3 = request.POST.get("price3") or None
            movies_id = request.POST.get("movie")
            rows = request.POST.get("rows")
            columns = request.POST.get("columns")
            existing_show_times = request.POST.getlist("existing_show_times[]")
            movies = Movies.objects.get(id=movies_id)

            # Update existing screen details
            screen.theatre = theatre
            screen.name = name
            screen.price1 = price1
            screen.price2 = price2
            screen.price3 = price3
            screen.movies = movies
            screen.total_seat_rows = rows
            screen.total_seat_columns = columns

            # Remove deselected show_times from the screen
            deselected_show_times = screen.show_times.exclude(
                id__in=existing_show_times
            )
            for show_time in deselected_show_times:
                screen.show_times.remove(show_time)

            # Add the selected show_times to the screen
            for show_time_id in existing_show_times:
                show_time = Show_Time.objects.get(id=show_time_id)
                screen.show_times.add(show_time)

            # Create new show_times and add them to the screen
            new_show_times = request.POST.getlist("show_times[]")
            for show_time_str in new_show_times:
                show_time_obj = Show_Time.objects.create(time=show_time_str)
                screen.show_times.add(show_time_obj)

            return redirect("theatre_screen")
        else:
            movies = Movies.objects.all()
            selected_show_times = screen.show_times.all()
            available_show_times = Show_Time.objects.exclude(
                id__in=selected_show_times.values_list("id", flat=True)
            )
            return render(
                request,
                "theatre/update_screen.html",
                {
                    "screens": screen,
                    "movies": movies,
                    "available_show_times": available_show_times,
                    "selected_show_times": selected_show_times,
                },
            )
    else:
        return redirect("home")


def delete_screen(request, id):
    if "user" in request.session:
        return redirect("home")
    if "admin" in request.session:
        return redirect("admin_home")
    if "theatre" in request.session:
        screens = Screen.objects.get(id=id)
        if request.method == "POST":
            screens.delete()
            return redirect("theatre_screen")
        else:
            return render(request, "theatre/delete_screen.html")
    else:
        return redirect("home")


def theatre_side_booking(request):
    bookings = BookedSeat.objects.filter(theatre=request.user).order_by("-id")
    theatre_sale_report = Theatre_Sale_Report.objects.filter(
        name=request.user
    ).order_by("-id")

    if request.method == "GET":
        # Filter by date range
        start_date = request.GET.get("start_date")
        end_date = request.GET.get("end_date")

        print("start_date", start_date, "end_date", end_date)
        if start_date and end_date:
            # Convert the start_date and end_date strings to datetime objects
            start_date = date.fromisoformat(start_date)
            end_date = date.fromisoformat(end_date)
        else:
            # Set default values for start_date and end_date
            today = date.today()
            start_date = today - timedelta(days=7)  # Default start_date is 7 days ago
            end_date = today
        if start_date and end_date:
            theatre_sale_report = Theatre_Sale_Report.objects.filter(
                name=request.user,
                booking__booked_date__gte=start_date,
                booking__booked_date__lte=end_date,
            ).order_by("-id")
    items_per_page = 13
    paginator = Paginator(theatre_sale_report, items_per_page)
    page_number = request.GET.get("page")
    page = paginator.get_page(page_number)
    context = {
        "theatre_sale_report": theatre_sale_report,
        "start_date": start_date,
        "end_date": end_date,
        "page": page,
    }
    print("start_date", start_date, "end_date", end_date)
    return render(request, "theatre/theatre_side_booking.html", context)


def theatre_generate_excel(request):
    if request.method == "GET":
        start_date = request.GET.get("start_date")
        end_date = request.GET.get("end_date")
        print("start_date:", start_date, "end_date", end_date)

        # Apply the filters to the queryset
        theatre_sale_report = Theatre_Sale_Report.objects.filter(name=request.user)

        if start_date and end_date:
            # Convert the start_date and end_date strings to datetime objects
            start_date = datetime.strptime(start_date, "%B %d, %Y").date()
            end_date = datetime.strptime(end_date, "%B %d, %Y").date()

            theatre_sale_report = theatre_sale_report.filter(
                booking__booked_date__range=[start_date, end_date]
            )

        # Create a new Excel workbook
        workbook = openpyxl.Workbook()
        sheet = workbook.active

        # Write headers
        headers = [
            "Booking ID",
            "Username",
            "Movie",
            "Screen",
            "Seats",
            "Show",
            "Date",
            "Admin Sale",
            "Payment ID",
            "Theatre Sale",
            "Booked",
        ]
        for col_num, header in enumerate(headers, 1):
            column_letter = get_column_letter(col_num)
            sheet[column_letter + "1"] = header
            sheet[column_letter + "1"].font = Font(bold=True)

        # Write data rows
        row_num = 2
        for sale_report in theatre_sale_report:
            booking = sale_report.booking

            row = [
                str(booking.booking_id),
                booking.user,
                booking.movie,
                booking.screen,
                booking.booked_seats if booking.booked_seats else "Cancelled",
                booking.show_time,
                booking.date,
                sale_report.theatre_earnings,
                booking.payment_id,
                sale_report.theatre_earnings,
                f'{booking.booked_date.strftime("%Y-%m-%d")}',
            ]

            for col_num, cell_value in enumerate(row, 1):
                column_letter = get_column_letter(col_num)
                sheet[column_letter + str(row_num)] = cell_value

            row_num += 1

        # Add total sums row
        total_sum_row = [
            "Total Theatre Sale:"
            + str(
                theatre_sale_report.aggregate(theatre_sale_sum=Sum("theatre_earnings"))[
                    "theatre_sale_sum"
                ]
            )
            + "/-"
        ]

        for col_num, cell_value in enumerate(total_sum_row, 1):
            column_letter = get_column_letter(col_num)
            sheet[column_letter + str(row_num)] = cell_value
            sheet[column_letter + str(row_num)].font = Font(bold=True)

        # Save the workbook and return it as a response
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = "attachment; filename=admin_sale_report.xlsx"
        workbook.save(response)

        return response


def cancellation_requests(request):
    user = UserProfile.objects.get(user=request.user)
    cancellation_requests = BookingCancellationRequest.objects.filter(
        theatre=user, status="pending"
    ).order_by("-id")
    return render(
        request,
        "theatre/cancellation_requests.html",
        {"cancellation_requests": cancellation_requests},
    )


def update_cancellation_request(request):
    if request.method == "POST":
        request_id = request.POST.get("request_id")
        status = request.POST.get("status")
        cancellation_request = BookingCancellationRequest.objects.get(id=request_id)
        cancellation_request.status = status
        cancellation_request.save()
        print("booking_id", cancellation_request.booking.id)
        print("status:", status)

        if status == "accepted":
            booking_id = cancellation_request.booking.id
            booked_seat = BookedSeat.objects.get(id=booking_id)
            # Retrieve the original booking details
            original_theatre_price = booked_seat.price
            original_theatre_earnings = original_theatre_price * Decimal("0.85")
            original_admin_earnings = original_theatre_price * Decimal("0.15")

            # Update the booked seats
            booked_seat.booked_seats = ""
            booked_seat.save()

            # Update theatre sale report
            theatre_sale_report = Theatre_Sale_Report.objects.get(booking=booked_seat)
            theatre_sale_report.theatre_earnings -= original_theatre_earnings
            theatre_sale_report.save()

            # Update admin sale report
            admin_sale_report = Admin_Sale_Report.objects.get(
                theatre_sale_report=theatre_sale_report
            )
            admin_sale_report.admin_earnings -= original_admin_earnings
            admin_sale_report.save()

            try:
                client = razorpay.Client(auth=(KEY, SECRET))
                paymentId = cancellation_request.booking.payment_id
                refund = client.payment.refund(
                    paymentId,
                    {
                        "amount": int(cancellation_request.booking.price) * 100,
                        "speed": "normal",
                        "notes": {
                            "notes_key_1": "Beam me up Scotty.",
                            "notes_key_2": "Engage",
                        },
                        "receipt": f"Receipt No. {booking_id}",
                    },
                )

                if refund["status"] == "processed":
                    # Refund successful
                    # Add your code here for any additional actions, such as updating the booking or notifying the user
                    print("Refund response:", json.dumps(refund, indent=4))
                    print("Refund processed successfully")
                else:
                    # Refund failed
                    # Handle the refund failure scenario
                    print("Refund response:", json.dumps(refund, indent=4))
                    print("Refund failed")
            except Exception as e:
                print("Error occurred during refund:", str(e))

        return JsonResponse({"status": "success"})
    return JsonResponse({"status": "error"})


def theatre_profile(request):
    user = User.objects.get(username=request.user)
    user_profile = UserProfile.objects.get(user=user)
    if request.method == "POST":
        form = UserProfileForm(request.POST, request.FILES, instance=user_profile)
        if form.is_valid():
            first_name = form.cleaned_data["first_name"]
            last_name = form.cleaned_data["last_name"]
            profile_pic = form.cleaned_data["profile_image"]
            username = form.cleaned_data["username"]
            email = form.cleaned_data["email"]
            phone = form.cleaned_data["phone"]
            address = form.cleaned_data["address"]
            location = form.cleaned_data["location"]

            # Update the user details
            user.first_name = first_name
            user.last_name = last_name
            user.username = username
            user.email = email
            user.save()

            # Update the user profile details
            user_profile.profile_image = profile_pic
            user_profile.phone = phone
            user_profile.address = address
            user_profile.location = location
            user_profile.save()
            return redirect("theatre_profile")
    else:
        user_data = {
            "first_name": user.first_name,
            "last_name": user.last_name,
            "username": user.username,
            "email": user.email,
        }
        form = UserProfileForm(instance=user_profile, initial=user_data)
        return render(
            request,
            "theatre/theatre_profile.html",
            {"form": form, "user_profile": user_profile},
        )


def theatre_logout(request):
    if "theatre" in request.session:
        request.session.flush()
    logout(request)
    return redirect("home")
