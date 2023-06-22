import calendar
from django.shortcuts import render, redirect
from django.contrib import auth
from .models import *
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from theatre.models import *
from home.models import *
from users.models import *
from home.forms import *
from datetime import date, timedelta
from django.db.models import Sum
from django.utils import timezone
import decimal, uuid
from datetime import date, timedelta, datetime
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
import openpyxl
from django.db.models.functions import ExtractMonth
from django.core.paginator import Paginator

# Create your views here.


def admin_home(request):
    if "user" in request.session:
        return redirect("home")
    if "theatre" in request.session:
        return redirect("theatre_home")
    if "admin" in request.session:
        # Get today's date
        today = date.today()

        # Get the start and end dates for the current month
        start_of_month = today.replace(day=1)
        end_of_month = start_of_month.replace(
            month=start_of_month.month + 1
        ) - timedelta(days=1)

        theatres = UserProfile.objects.filter(is_theatre=True)

        # Create an empty dictionary to store the sale reports for each theatre
        theatre_sale_reports = {}

        # Iterate over the theatres
        for theatre in theatres:
            # Get the sale reports for the theatre for the current month
            theatre_reports = Admin_Sale_Report.objects.filter(
                theatre_sale_report__booking__booked_date__range=(
                    start_of_month,
                    end_of_month,
                ),
                theatre_sale_report__booking__theatre=theatre,
            )

            # Add the theatre's sale reports to the dictionary
            theatre_sale_reports[theatre.user] = theatre_reports

        sales_data = []
        revenue_data = []
        days_of_week = []

        for i in range(7):
            # Get the date for the corresponding day of the week
            day = today - timedelta(days=today.weekday()) + timedelta(days=i)
            days_of_week.append(day.strftime("%a"))

            day_total_price = sum(
                report.theatre_sale_report.booking.price
                for report in Admin_Sale_Report.objects.filter(
                    theatre_sale_report__date_added=day
                )
            )
            day_total_revenue = sum(
                report.admin_earnings
                for report in Admin_Sale_Report.objects.filter(
                    theatre_sale_report__date_added=day
                )
            )

            today_total_price = sum(
                report.theatre_sale_report.booking.price
                for report in Admin_Sale_Report.objects.filter(
                    theatre_sale_report__date_added=today
                )
            )
            today_total_revenue = sum(
                report.admin_earnings
                for report in Admin_Sale_Report.objects.filter(
                    theatre_sale_report__date_added=today
                )
            )

            sales_data.append(day_total_price)
            revenue_data.append(day_total_revenue)

        sale_reports = Admin_Sale_Report.objects.all()

        month_total_price = (
            Admin_Sale_Report.objects.filter(
                theatre_sale_report__date_added__range=(start_of_month, end_of_month)
            ).aggregate(total_price=Sum("theatre_sale_report__booking__price"))[
                "total_price"
            ]
            or 0.00
        )

        month_total_revenue = (
            Admin_Sale_Report.objects.filter(
                theatre_sale_report__date_added__range=(start_of_month, end_of_month)
            ).aggregate(total_price=Sum("admin_earnings"))["total_price"]
            or 0.00
        )

        print(
            "sales_data:",
            sales_data,
            "revenue_data:",
            revenue_data,
            "days_of_week:",
            days_of_week,
        )

        revenue_data = [
            float(value) if isinstance(value, decimal.Decimal) else value
            for value in revenue_data
        ]
        context = {
            "sale_reports": sale_reports,
            "theatre_sale_reports": theatre_sale_reports,
            "day_total_price": day_total_price,
            "month_total_price": month_total_price,
            "day_total_revenue": day_total_revenue,
            "month_total_revenue": month_total_revenue,
            "sales_data": sales_data,
            "revenue_data": revenue_data,
            "days_of_week": days_of_week,
            "today_total_price": today_total_price,
            "today_total_revenue": today_total_revenue,
        }
        print(
            "today_total_price:",
            day_total_price,
            "today_total_revenue:",
            day_total_revenue,
        )
        return render(request, "admin_panel/admin_home.html", context)
    else:
        return redirect("home")


def admin_profile(request):
    user = User.objects.get(username=request.user)
    if request.method == "POST":
        form = AdminProfileForm(request.POST, request.FILES, instance=user)
        if form.is_valid():
            first_name = form.cleaned_data["first_name"]
            last_name = form.cleaned_data["last_name"]
            username = form.cleaned_data["username"]
            email = form.cleaned_data["email"]

            # Update the user details
            user.first_name = first_name
            user.last_name = last_name
            user.username = username
            user.email = email
            user.save()
            return redirect("admin_profile")
    else:
        user_data = {
            "first_name": user.first_name,
            "last_name": user.last_name,
            "username": user.username,
            "email": user.email,
        }
        form = AdminProfileForm(instance=user, initial=user_data)
        return render(
            request, "admin_panel/admin_profile.html", {"form": form, "user": user}
        )


def admin_movies(request):
    if "user" in request.session:
        return redirect("home")
    if "theatre" in request.session:
        return redirect("theatre_home")
    if "admin" in request.session:
        movies = Movies.objects.all()
        items_per_page = 4
        paginator = Paginator(movies, items_per_page)
        page_number = request.GET.get("page")
        page = paginator.get_page(page_number)
        return render(
            request, "admin_panel/admin_movies.html", {"movies": movies, "page": page}
        )
    else:
        return redirect("home")


def add_movie(request):
    if "user" in request.session:
        return redirect("home")
    if "theatre" in request.session:
        return redirect("theatre_home")
    if "admin" in request.session:
        if request.method == "POST":
            title = request.POST.get("title")
            description = request.POST.get("description") or None
            cast = request.POST.get("cast")
            writers = request.POST.get("writers")
            director = request.POST.get("director")
            year = request.POST.get("year")
            genre = request.POST.get("genre") or None
            image = request.FILES.get("image")
            poster = request.FILES.get("poster")
            release_date = request.POST.get("release_date")
            language_id = request.POST.get("language")
            trailer = request.POST.get("trailer")
            runtime = request.POST.get("runtime") or None
            language = All_Languages.objects.get(id=language_id)
            movie = Movies.objects.create(
                title=title,
                description=description,
                genre=genre,
                cast=cast,
                director=director,
                writers=writers,
                year=year,
                image=image,
                poster=poster,
                release_date=release_date,
                language=language,
                trailer=trailer,
                runtime=runtime,
            )
            movie.save()
            return redirect("admin_movies")
        else:
            languages = All_Languages.objects.all()
            return render(
                request, "admin_panel/add_movie.html", {"languages": languages}
            )
    else:
        return redirect("home")


def update_movie(request, id):
    if "user" in request.session:
        return redirect("home")
    if "theatre" in request.session:
        return redirect("theatre_home")
    if "admin" in request.session:
        movies = Movies.objects.get(id=id)
        if request.method == "POST":
            title = request.POST.get("title")
            description = request.POST.get("description") or None
            cast = request.POST.get("cast")
            writers = request.POST.get("writers")
            director = request.POST.get("director")
            year = request.POST.get("year")
            genre = request.POST.get("genre") or None
            image = request.FILES.get("image")
            poster = request.FILES.get("poster")
            release_date = request.POST.get("release_date")
            language_id = request.POST.get("language")
            trailer = request.POST.get("trailer")
            runtime = request.POST.get("runtime") or None
            language = All_Languages.objects.get(id=language_id)
            movie = Movies(
                id=id,
                title=title,
                description=description,
                genre=genre,
                cast=cast,
                writers=writers,
                director=director,
                year=year,
                image=image,
                poster=poster,
                release_date=release_date,
                language=language,
                trailer=trailer,
                runtime=runtime,
            )
            movie.save()
            return redirect("admin_movies")
        else:
            languages = All_Languages.objects.all()
            print("image url", movies.image)
            return render(
                request,
                "admin_panel/update_movie.html",
                {"movies": movies, "languages": languages},
            )
    else:
        return redirect("home")


def delete_movie(request, id):
    if "user" in request.session:
        return redirect("home")
    if "theatre" in request.session:
        return redirect("theatre_home")
    if "admin" in request.session:
        movies = Movies.objects.get(id=id)
        if request.method == "POST":
            movies.delete()
            return redirect("admin_movies")
        else:
            return render(request, "admin_panel/delete_movie.html")
    else:
        return redirect("home")


def admin_theatres(request):
    if "user" in request.session:
        return redirect("home")
    if "theatre" in request.session:
        return redirect("theatre_home")
    if "admin" in request.session:
        theatre = UserProfile.objects.filter(is_theatre=True).order_by("id")
        items_per_page = 12
        paginator = Paginator(theatre, items_per_page)
        page_number = request.GET.get("page")
        page = paginator.get_page(page_number)
        return render(
            request,
            "admin_panel/admin_theatres.html",
            {"theatres": theatre, "page": page},
        )
    else:
        return redirect("home")


def block_theatre(request, id):
    if request.method == "POST":
        user = User.objects.get(id=id)
        user.is_active = False
        user.save()
        return redirect("admin_theatres")


def unblock_theatre(request, id):
    if request.method == "POST":
        user = User.objects.get(id=id)
        user.is_active = True
        user.save()
        return redirect("admin_theatres")
    else:
        return redirect("admin_theatres")


def admin_users(request):
    if "user" in request.session:
        return redirect("home")
    if "theatre" in request.session:
        return redirect("theatre_home")
    if "admin" in request.session:
        users = User.objects.filter(userprofile__is_theatre=False).order_by("id")
        items_per_page = 12
        paginator = Paginator(users, items_per_page)
        page_number = request.GET.get("page")
        page = paginator.get_page(page_number)
        return render(
            request, "admin_panel/admin_users.html", {"users": users, "page": page}
        )
    else:
        return redirect("home")


def block_user(request, id):
    if request.method == "POST":
        user = User.objects.get(id=id)
        user.is_active = False
        user.save()
        return redirect("admin_users")


def unblock_user(request, id):
    if request.method == "POST":
        user = User.objects.get(id=id)
        user.is_active = True
        user.save()
        return redirect("admin_users")


def admin_side_booking(request):
    bookings = BookedSeat.objects.all().order_by("-id")
    admin_sale_report = Admin_Sale_Report.objects.all().order_by("-id")
    theatres = UserProfile.objects.filter(is_theatre=True)

    if request.method == "GET":
        selected_theatre = request.GET.get("theatre_filter")
        # Filter by date range
        start_date = request.GET.get("start_date")
        end_date = request.GET.get("end_date")

        print(
            "selected_theatre",
            selected_theatre,
            "start_date",
            start_date,
            "end_date",
            end_date,
        )
        if start_date and end_date:
            # Convert the start_date and end_date strings to datetime objects
            start_date = datetime.strptime(start_date, "%Y-%m-%d").date()
            end_date = datetime.strptime(end_date, "%Y-%m-%d").date()
        else:
            # Set default values for start_date and end_date
            today = datetime.now().date()
            start_date = today - timedelta(days=7)  # Default start_date is 7 days ago
            end_date = today

        if selected_theatre:
            admin_sale_report = admin_sale_report.filter(
                theatre_sale_report__booking__theatre=selected_theatre
            )
        if start_date and end_date:
            admin_sale_report = admin_sale_report.filter(
                theatre_sale_report__booking__booked_date__gte=start_date,
                theatre_sale_report__booking__booked_date__lte=end_date,
            )
    items_per_page = 12
    paginator = Paginator(admin_sale_report, items_per_page)
    page_number = request.GET.get("page")
    page = paginator.get_page(page_number)
    context = {
        "admin_sale_report": admin_sale_report,
        "theatres": theatres,
        "selected_theatre": selected_theatre,
        "start_date": start_date,
        "end_date": end_date,
        "page": page,
    }
    print("start_date", start_date, "end_date", end_date)
    return render(request, "admin_panel/admin_side_booking.html", context)


def generate_excel(request):
    # Retrieve the filter parameters from the request
    if request.method == "GET":
        selected_theatre = request.GET.get("theatre_filter")
        start_date = request.GET.get("start_date")
        end_date = request.GET.get("end_date")
        print(
            "selected_theatre:",
            selected_theatre,
            "start_date:",
            start_date,
            "end_date",
            end_date,
        )

        # Apply the filters to the queryset
        admin_sale_report = Admin_Sale_Report.objects.all()

        if selected_theatre:
            admin_sale_report = admin_sale_report.filter(
                theatre_sale_report__booking__theatre=selected_theatre
            )

        if start_date and end_date:
            # Convert the start_date and end_date strings to datetime objects
            start_date = datetime.strptime(start_date, "%B %d, %Y").date()
            end_date = datetime.strptime(end_date, "%B %d, %Y").date()

            admin_sale_report = admin_sale_report.filter(
                theatre_sale_report__booking__booked_date__range=[start_date, end_date]
            )

        # Create a new Excel workbook
        workbook = openpyxl.Workbook()
        sheet = workbook.active

        # Write headers
        headers = [
            "Booking ID",
            "Username",
            "Movie",
            "Theatre",
            "Screen",
            "Seats",
            "Show",
            "Date",
            "Admin Sale",
            "Payment ID",
            "Theatre Sale",
            "Booked",
        ]
        for col_num, header in enumerate(headers, 1):
            column_letter = get_column_letter(col_num)
            sheet[column_letter + "1"] = header
            sheet[column_letter + "1"].font = Font(bold=True)

        # Write data rows
        row_num = 2
        for sale_report in admin_sale_report:
            booking = sale_report.theatre_sale_report.booking

            row = [
                str(booking.booking_id),
                booking.user,
                booking.movie,
                booking.theatre,
                booking.screen,
                booking.booked_seats if booking.booked_seats else "Cancelled",
                booking.show_time,
                booking.date,
                sale_report.admin_earnings,
                booking.payment_id,
                sale_report.theatre_sale_report.theatre_earnings,
                f'{booking.booked_date.strftime("%Y-%m-%d")}',
            ]

            for col_num, cell_value in enumerate(row, 1):
                column_letter = get_column_letter(col_num)
                sheet[column_letter + str(row_num)] = cell_value

            row_num += 1

        # Add total sums row
        total_sum_row = [
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "Total Admin Sale: "
            + str(
                admin_sale_report.aggregate(admin_sale_sum=Sum("admin_earnings"))[
                    "admin_sale_sum"
                ]
            )
            + "/-",
            "",
            "",
            "Total Theatre Sale:"
            + str(
                admin_sale_report.aggregate(
                    theatre_sale_sum=Sum("theatre_sale_report__theatre_earnings")
                )["theatre_sale_sum"]
            )
            + "/-",
        ]

        for col_num, cell_value in enumerate(total_sum_row, 1):
            column_letter = get_column_letter(col_num)
            sheet[column_letter + str(row_num)] = cell_value
            sheet[column_letter + str(row_num)].font = Font(bold=True)

        # Save the workbook and return it as a response
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = "attachment; filename=admin_sale_report.xlsx"
        workbook.save(response)

        return response


def admin_cancellation_requests(request):
    cancellation_requests = BookingCancellationRequest.objects.all().order_by("-id")
    return render(
        request,
        "admin_panel/admin_cancellation_requests.html",
        {"cancellation_requests": cancellation_requests},
    )


def admin_logout(request):
    if "admin" in request.session:
        request.session.flush()
    logout(request)
    return redirect("home")
